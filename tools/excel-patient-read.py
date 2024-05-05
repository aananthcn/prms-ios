import sys
import openpyxl
import json

def read_treatment_sheet(workbook, patient_id):
    try:
        # Select the sheet with the given patient_id
        sheet = workbook[patient_id[0:31]]
        
        # Read treatment details from columns A to E
        treatments = []
        for row in sheet.iter_rows(values_only=True):
            # Assuming the data starts from the second row
            tid, date, complaint, prescription, doctor = row[:5]

            if prescription == None:
                prescription = ""
            if complaint == None:
                complaint = ""
            if doctor == None:
                doctor = "Dr. Jegadish"
            if "." in date:
                date = date.replace('.', '-')

            treatments.append({
                "id": str(tid),
                "dateString": date,
                "dateFormat":"yyyy-MM-dd",
                "complaint": complaint,
                "prescription": prescription,
                "doctor": {
                    "phone": "",
                    "id": doctor,
                    "name": doctor
                }
            })
        
        return treatments
    except Exception as e:
        print(f"Error reading treatment sheet for patient {patient_id}: {e}")
        return []

def read_excel(filename):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(filename)
        
        # Select the sheet named "patientlist"
        patientlist_sheet = workbook['patientlist']
        
        # Read data from columns B to J (ignoring column A)
        data = []
        for row in patientlist_sheet.iter_rows(min_row=2, values_only=True):
            name, phone, email, address, gender, patient_id, year, month, age = row[1:10]

            # Check if email is null
            if email is None:
                email = ""
            if address is None:
                address = ""
            if phone is None:
                phone = ""
            if gender is None:
                gender = ""

            patient_data = {
                "name": name,
                "phone": phone,
                "email": email,
                "address": address,
                "gender": gender,
                "id": patient_id,
                "year": year,
                "month": month,
                "age": age,
                "treatments": read_treatment_sheet(workbook, patient_id)
            }

            # Add the line here to remove the first entry from treatments
            patient_data["treatments"] = patient_data["treatments"][1:]

            data.append(patient_data)
            print(patient_data)
        
        return data
    except Exception as e:
        print("Error:", e)
        return None

if __name__ == "__main__":
    # Check if filename is provided as a command line argument
    if len(sys.argv) != 2:
        print("Usage: python program.py filename.xlsx")
        sys.exit(1)

    filename = sys.argv[1]
    data = read_excel(filename)
    
    if data:
        print("Data read successfully from", filename)
        # Write the data to a JSON file
        output_filename = filename.split('.')[0] + '_output.json'
        with open(output_filename, 'w') as outfile:
            json.dump(data, outfile, indent=4)
        print("JSON data written to", output_filename)
    else:
        print("Failed to read data from", filename)
